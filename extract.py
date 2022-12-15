from openpyxl import load_workbook


FILE_PATH = "D:\\VSProject\\pythonWorkspace\\realestate_main\\realestate\\시세트래킹.xlsx"
PRIFIX_PATH = "D:\\VSProject\\pythonWorkspace\\realestate_main\\realestate\\시세표 "

class Extraction:
    def __init__(self) -> None:
        self.load_wb = load_workbook(FILE_PATH, data_only=True)
        
        # 시트 이름으로 불러오기 
        load_ws = self.load_wb['수도권']
        #load_ws['A3'].value = "abc"
        print(load_ws['B3'].value)

        self.price_load_ws = None

        for row in load_ws.rows:
            if row[0].value == None:
                continue
    
            cell_value = row[0].value

            if "서울시" in cell_value or "경기도" in cell_value:
                path_str = PRIFIX_PATH + cell_value + ".xlsm"
                self.price_load_wb = load_workbook(path_str, data_only=True)
                #print(cell_value)
            elif "동" in cell_value and len(cell_value) == 3:
                #print(cell_value)
                self.price_load_ws = self.price_load_wb[cell_value]
            else:
                if self.price_load_ws != None:
                    for row in self.price_load_ws.rows:
                        print(row[0].value)

            

        #load_wb.save(FILE_PATH)

    

"""
    @brief      main entry
"""
if __name__ == "__main__":
    extract = Extraction()



